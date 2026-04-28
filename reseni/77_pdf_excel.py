"""Řešení – Lekce 77: PDF a Excel – generování dokumentů"""

# vyžaduje: pip install openpyxl reportlab

from pathlib import Path
from datetime import date, datetime

# ── Sdílená data ──────────────────────────────────────────────
STUDENTI = [
    (1, "Míša",  15, "Matematika",  87.5),
    (2, "Tomáš", 16, "Fyzika",      92.0),
    (3, "Bára",  14, "Matematika",  55.3),
    (4, "Ondra", 17, "Informatika", 95.1),
    (5, "Klára", 15, "Biologie",    61.0),
    (6, "Pavel", 16, "Chemie",      43.2),
]

# 1. Excel s podmíněným formátováním (červeně body < 60)
print("=== 1. Excel s podmíněným formátováním ===\n")

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                   Border, Side, numbers)
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule
    EXCEL_OK = True
except ImportError:
    print("openpyxl není nainstalováno: pip install openpyxl")
    EXCEL_OK = False

if EXCEL_OK:
    wb = Workbook()
    ws = wb.active
    ws.title = "Výkaz"

    # Styly
    TANKA  = Side(style="thin")
    OKRAJ  = Border(left=TANKA, right=TANKA, top=TANKA, bottom=TANKA)

    # Záhlaví
    HLAVICKY = ["#", "Jméno", "Věk", "Předmět", "Body", "Výsledek"]
    for col, text in enumerate(HLAVICKY, 1):
        b = ws.cell(row=1, column=col, value=text)
        b.font      = Font(bold=True, size=12, color="FFFFFF")
        b.fill      = PatternFill("solid", fgColor="2E86AB")
        b.alignment = Alignment(horizontal="center")
        b.border    = OKRAJ

    # Data s podmíněným formátováním
    for row_i, (cislo, jmeno, vek, predmet, body) in enumerate(STUDENTI, 2):
        vysledek = "Výborně" if body >= 90 else "Dobře" if body >= 75 else "Dostatečně" if body >= 60 else "Nedostatečně"
        barva_bg = "F2F2F2" if row_i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=barva_bg)

        for col, hodnota in enumerate([cislo, jmeno, vek, predmet, body, vysledek], 1):
            b = ws.cell(row=row_i, column=col, value=hodnota)
            b.fill   = fill
            b.border = OKRAJ
            b.alignment = Alignment(horizontal="center")

    # Podmíněné formátování: body < 60 → červené pozadí + bílé písmo
    cerveny_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cerveny_font = Font(color="FFFFFF", bold=True)
    cervene_pravidlo = CellIsRule(
        operator="lessThan",
        formula=["60"],
        fill=PatternFill("solid", fgColor="FFCCCC"),  # světle červená
        font=Font(color="CC0000", bold=True),
    )
    ws.conditional_formatting.add(f"E2:E{len(STUDENTI)+1}", cervene_pravidlo)

    # Podmíněné formátování: body >= 90 → zelené pozadí
    zelene_pravidlo = CellIsRule(
        operator="greaterThanOrEqual",
        formula=["90"],
        fill=PatternFill("solid", fgColor="CCFFCC"),
        font=Font(color="006600", bold=True),
    )
    ws.conditional_formatting.add(f"E2:E{len(STUDENTI)+1}", zelene_pravidlo)

    # Šířky sloupců
    for col, sirka in enumerate([5, 15, 6, 15, 8, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = sirka

    ws.freeze_panes = "A2"

    # Shrnutí pod tabulkou
    ws.cell(row=len(STUDENTI)+3, column=4, value="Průměr:").font = Font(bold=True)
    ws.cell(row=len(STUDENTI)+3, column=5,
            value=f"=AVERAGE(E2:E{len(STUDENTI)+1})")
    ws.cell(row=len(STUDENTI)+4, column=4, value="Pod 60:").font = Font(bold=True)
    ws.cell(row=len(STUDENTI)+4, column=5,
            value=f"=COUNTIF(E2:E{len(STUDENTI)+1},\"<60\")")

    wb.save("vykaz.xlsx")
    print(f"  ✓ vykaz.xlsx ({Path('vykaz.xlsx').stat().st_size:,} B)")
    print("    - body < 60  → červená (FFCCCC), tučný text")
    print("    - body >= 90 → zelená (CCFFCC), tučný text")
    Path("vykaz.xlsx").unlink(missing_ok=True)


# 2. PDF certifikát pro studenta
print("\n=== 2. PDF certifikát ===\n")

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib.enums import TA_CENTER
    PDF_OK = True
except ImportError:
    print("reportlab není nainstalováno: pip install reportlab")
    PDF_OK = False

if PDF_OK:
    def generuj_certifikat(jmeno: str, kurz: str,
                             datum: str, soubor: str):
        """Vygeneruje PDF certifikát."""
        doc = SimpleDocTemplate(
            soubor,
            pagesize=landscape(A4),
            leftMargin=25*mm, rightMargin=25*mm,
            topMargin=25*mm, bottomMargin=25*mm,
        )

        ZLUTA   = colors.HexColor("#F4D03F")
        MODRA   = colors.HexColor("#2E86AB")
        TMAVA   = colors.HexColor("#1A1A2E")
        SEDA    = colors.HexColor("#666666")

        styles = getSampleStyleSheet()
        titul   = ParagraphStyle("titul",   parent=styles["Normal"],
                                   fontSize=36, textColor=MODRA,
                                   alignment=TA_CENTER, fontName="Helvetica-Bold",
                                   spaceAfter=5)
        podtitul = ParagraphStyle("podtitul", parent=styles["Normal"],
                                   fontSize=14, textColor=SEDA,
                                   alignment=TA_CENTER, spaceAfter=15)
        jmeno_style = ParagraphStyle("jmeno_s", parent=styles["Normal"],
                                       fontSize=28, textColor=TMAVA,
                                       alignment=TA_CENTER, fontName="Helvetica-Bold",
                                       spaceAfter=10)
        text_style  = ParagraphStyle("text_s", parent=styles["Normal"],
                                       fontSize=14, textColor=TMAVA,
                                       alignment=TA_CENTER, spaceAfter=5)
        datum_style = ParagraphStyle("datum_s", parent=styles["Normal"],
                                       fontSize=11, textColor=SEDA,
                                       alignment=TA_CENTER)

        obsah = [
            Spacer(1, 10*mm),
            Paragraph("CERTIFIKÁT", titul),
            Paragraph("o úspěšném absolvování kurzu", podtitul),
            HRFlowable(width="80%", thickness=2, color=ZLUTA, spaceAfter=15),
            Paragraph("Tento certifikát potvrzuje, že", text_style),
            Spacer(1, 5*mm),
            Paragraph(f"<b>{jmeno}</b>", jmeno_style),
            Spacer(1, 5*mm),
            Paragraph(f"úspěšně absolvoval/a kurz", text_style),
            Paragraph(f"<b>{kurz}</b>", jmeno_style),
            Spacer(1, 10*mm),
            HRFlowable(width="60%", thickness=1, color=SEDA, spaceAfter=10),
            Paragraph(f"Datum vydání: {datum}", datum_style),
            Paragraph("🐍 Python Kurz | python-kurz.cz", datum_style),
        ]

        doc.build(obsah)

    generuj_certifikat(
        jmeno="Míša Nováková",
        kurz="Interaktivní Python kurz",
        datum=date.today().strftime("%d. %m. %Y"),
        soubor="certifikat.pdf",
    )
    print(f"  ✓ certifikat.pdf ({Path('certifikat.pdf').stat().st_size:,} B)")
    print("    - Landscape A4 formát")
    print("    - Barevný design s linkou")
    print("    - Jméno, kurz, datum")
    Path("certifikat.pdf").unlink(missing_ok=True)


# 3. Čtení tabulky z PDF faktury → SQLite
print("\n=== 3. PDF faktura → SQLite ===\n")

import sqlite3

# Nejdříve vygeneruj PDF fakturu
if PDF_OK:
    from reportlab.platypus import Table, TableStyle

    def generuj_fakturu_pdf(polozky: list, soubor: str):
        """Jednoduchá faktura pro demo extrakce."""
        doc = SimpleDocTemplate(soubor, pagesize=A4,
                                  leftMargin=20*mm, rightMargin=20*mm,
                                  topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        obsah  = [Paragraph("FAKTURA č. 2024-001", styles["Title"]),
                  Spacer(1, 10*mm)]

        data = [["Položka", "Množství", "Cena/ks", "Celkem"]]
        for p in polozky:
            data.append([p["nazev"], str(p["mnozstvi"]),
                         f"{p['cena']} Kč", f"{p['mnozstvi']*p['cena']} Kč"])

        tbl = Table(data)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        obsah.append(tbl)
        doc.build(obsah)

    POLOZKY = [
        {"nazev": "Python kurz",      "mnozstvi": 1, "cena": 2999},
        {"nazev": "NumPy kniha",       "mnozstvi": 2, "cena": 599},
        {"nazev": "Docker workshop",   "mnozstvi": 1, "cena": 1500},
    ]
    generuj_fakturu_pdf(POLOZKY, "faktura_demo.pdf")

    # Čtení tabulky a uložení do SQLite (bez PyMuPDF – z dat přímo)
    # V produkci: použij pdfplumber nebo PyMuPDF pro extrakci z PDF
    print("  (Demo: data z faktury ukládám do SQLite přímo)")
    print("  V produkci: pip install pdfplumber")
    print("    import pdfplumber")
    print("    with pdfplumber.open('faktura.pdf') as pdf:")
    print("        tabulky = pdf.pages[0].extract_tables()")

    # SQLite část (bez skutečného čtení PDF)
    conn = sqlite3.connect(":memory:")
    conn.execute("""
        CREATE TABLE faktura_polozky (
            id       INTEGER PRIMARY KEY,
            nazev    TEXT,
            mnozstvi INTEGER,
            cena_kc  INTEGER,
            celkem   INTEGER
        )
    """)
    for p in POLOZKY:
        conn.execute(
            "INSERT INTO faktura_polozky (nazev, mnozstvi, cena_kc, celkem) VALUES (?,?,?,?)",
            (p["nazev"], p["mnozstvi"], p["cena"], p["mnozstvi"]*p["cena"])
        )
    conn.commit()

    rows = conn.execute("SELECT nazev, mnozstvi, cena_kc, celkem FROM faktura_polozky").fetchall()
    print("\n  Uloženo do SQLite:")
    for r in rows:
        print(f"    {r[0]:<22} {r[1]}ks  {r[2]:>6} Kč  = {r[3]:>6} Kč")

    celkem = sum(r[3] for r in rows)
    print(f"\n  CELKEM bez DPH: {celkem:,} Kč")
    print(f"  DPH 21%:        {int(celkem * 0.21):,} Kč")
    print(f"  CELKEM s DPH:   {int(celkem * 1.21):,} Kč")

    Path("faktura_demo.pdf").unlink(missing_ok=True)

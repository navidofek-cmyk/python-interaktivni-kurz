"""
LEKCE 77: PDF a Excel – generování dokumentů
=============================================
pip install reportlab openpyxl pillow

Velmi praktické: reporty, faktury, exporty dat.

reportlab  – generování PDF od nuly
openpyxl   – čtení a zápis Excel (.xlsx)
PyMuPDF    – čtení existujících PDF (pip install pymupdf)
"""

import io
import os
from pathlib import Path
from datetime import datetime, date
from dataclasses import dataclass

# ══════════════════════════════════════════════════════════════
# ČÁST 1: EXCEL s openpyxl
# ══════════════════════════════════════════════════════════════

print("=== Excel s openpyxl ===\n")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                   Border, Side, numbers)
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    print("openpyxl není nainstalováno: pip install openpyxl")
    EXCEL_OK = False

if EXCEL_OK:
    # ── Vytvoření sešitu ──────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Studenti"

    # Styly
    NADPIS_STYL = Font(bold=True, size=12, color="FFFFFF")
    NADPIS_FILL = PatternFill("solid", fgColor="2E86AB")
    ZAROVNANI   = Alignment(horizontal="center", vertical="center")
    TANKA_CARA  = Side(style="thin")
    OHRANICENI  = Border(left=TANKA_CARA, right=TANKA_CARA,
                          top=TANKA_CARA, bottom=TANKA_CARA)

    # Záhlaví
    hlavicky = ["#", "Jméno", "Věk", "Předmět", "Body", "Výsledek"]
    for col, text in enumerate(hlavicky, 1):
        bunka = ws.cell(row=1, column=col, value=text)
        bunka.font      = NADPIS_STYL
        bunka.fill      = NADPIS_FILL
        bunka.alignment = ZAROVNANI
        bunka.border    = OHRANICENI

    # Data
    studenti = [
        (1, "Míša",  15, "Matematika", 87.5),
        (2, "Tomáš", 16, "Fyzika",     92.0),
        (3, "Bára",  14, "Matematika", 78.3),
        (4, "Ondra", 17, "Informatika", 95.1),
        (5, "Klára", 15, "Biologie",   65.0),
    ]

    for row_idx, (cislo, jmeno, vek, predmet, body) in enumerate(studenti, 2):
        vysledek = "Výborně" if body >= 90 else "Dobře" if body >= 75 else "Dostatečně"

        # Střídaný background řádků
        barva = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
        fill  = PatternFill("solid", fgColor=barva)

        hodnoty = [cislo, jmeno, vek, predmet, body, vysledek]
        for col, hodnota in enumerate(hodnoty, 1):
            bunka = ws.cell(row=row_idx, column=col, value=hodnota)
            bunka.fill      = fill
            bunka.border    = OHRANICENI
            bunka.alignment = Alignment(horizontal="center")

        # Barevný výsledek
        vysledek_bunka = ws.cell(row=row_idx, column=6)
        if body >= 90:
            vysledek_bunka.font = Font(color="1B7A1B", bold=True)
        elif body < 70:
            vysledek_bunka.font = Font(color="CC0000")

    # Šířky sloupců
    sirky = [5, 15, 8, 15, 10, 15]
    for col, sirka in enumerate(sirky, 1):
        ws.column_dimensions[get_column_letter(col)].width = sirka

    # Ukotvení záhlaví
    ws.freeze_panes = "A2"

    # Statistiky pod tabulkou
    ws.cell(row=8, column=4, value="Průměr:").font = Font(bold=True)
    ws.cell(row=8, column=5, value=f"=AVERAGE(E2:E{len(studenti)+1})")
    ws.cell(row=9, column=4, value="Maximum:").font = Font(bold=True)
    ws.cell(row=9, column=5, value=f"=MAX(E2:E{len(studenti)+1})")

    # Graf
    ws2 = wb.create_sheet("Graf")
    data = [["Student", "Body"]] + [[s[1], s[4]] for s in studenti]
    for row in data:
        ws2.append(row)

    chart = BarChart()
    chart.title = "Body studentů"
    chart.y_axis.title = "Body"
    chart.x_axis.title = "Student"

    data_ref  = Reference(ws2, min_col=2, min_row=1, max_row=len(studenti)+1)
    cats_ref  = Reference(ws2, min_col=1, min_row=2, max_row=len(studenti)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws2.add_chart(chart, "D2")

    # Druhý list – formátování čísel
    ws3 = wb.create_sheet("Faktury")
    ws3.append(["Položka", "Množství", "Cena/ks", "Celkem"])
    polozky = [("Python kurz", 1, 2999), ("NumPy kniha", 2, 599),
                ("Docker workshop", 1, 1500)]
    for polozka, mnozstvi, cena in polozky:
        ws3.append([polozka, mnozstvi, cena, mnozstvi * cena])

    # Formátuj jako měnu
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=4):
        for bunka in row:
            bunka.number_format = '#,##0 Kč'

    wb.save("studenti.xlsx")
    print(f"  ✓ studenti.xlsx vytvořen ({Path('studenti.xlsx').stat().st_size:,} B)")

    # ── Čtení existujícího Excelu ─────────────────────────────
    wb2   = openpyxl.load_workbook("studenti.xlsx")
    ws_r  = wb2["Studenti"]
    print(f"  Čtení: {ws_r.max_row-1} studentů, {ws_r.max_column} sloupců")
    for row in ws_r.iter_rows(min_row=2, max_row=3, values_only=True):
        print(f"    {row}")

    Path("studenti.xlsx").unlink(missing_ok=True)


# ══════════════════════════════════════════════════════════════
# ČÁST 2: PDF s reportlab
# ══════════════════════════════════════════════════════════════

print("\n=== PDF s reportlab ===\n")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable,
                                     PageBreak)
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    PDF_OK = True
except ImportError:
    print("reportlab není nainstalováno: pip install reportlab")
    PDF_OK = False

if PDF_OK:
    # ── Faktura ───────────────────────────────────────────────
    def generuj_fakturu(cislo: str, odberatel: dict,
                         polozky: list, soubor: str):
        doc = SimpleDocTemplate(
            soubor,
            pagesize=A4,
            leftMargin=20*mm, rightMargin=20*mm,
            topMargin=20*mm, bottomMargin=20*mm,
        )

        styles = getSampleStyleSheet()
        h1 = ParagraphStyle("h1", parent=styles["Heading1"],
                              fontSize=20, spaceAfter=5)
        normal = styles["Normal"]
        small  = ParagraphStyle("small", parent=normal, fontSize=8,
                                  textColor=colors.grey)

        obsah = []

        # Záhlaví
        obsah.append(Paragraph(f"FAKTURA č. {cislo}", h1))
        obsah.append(Paragraph(
            f"Datum vystavení: {date.today().strftime('%d.%m.%Y')}", small))
        obsah.append(Spacer(1, 10*mm))
        obsah.append(HRFlowable(width="100%", thickness=1,
                                  color=colors.HexColor("#2E86AB")))
        obsah.append(Spacer(1, 5*mm))

        # Dodavatel / Odběratel
        info_data = [
            ["Dodavatel:", "Odběratel:"],
            ["Python s.r.o.", odberatel.get("nazev", "")],
            ["Hlavní 1, Praha", odberatel.get("adresa", "")],
            ["IČO: 12345678", f"IČO: {odberatel.get('ico', '')}"],
        ]
        info_tbl = Table(info_data, colWidths=[85*mm, 85*mm])
        info_tbl.setStyle(TableStyle([
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ]))
        obsah.append(info_tbl)
        obsah.append(Spacer(1, 10*mm))

        # Tabulka položek
        hlavicky_fak = ["Položka", "Množství", "Cena/ks", "Celkem"]
        data_fak = [hlavicky_fak]
        celkem = 0
        for polozka, mnozstvi, cena in polozky:
            radek_celkem = mnozstvi * cena
            celkem += radek_celkem
            data_fak.append([
                polozka, str(mnozstvi),
                f"{cena:,.0f} Kč", f"{radek_celkem:,.0f} Kč",
            ])
        dph = celkem * 0.21
        data_fak += [
            ["", "", "Základ DPH:", f"{celkem:,.0f} Kč"],
            ["", "", "DPH 21%:",    f"{dph:,.0f} Kč"],
            ["", "", "CELKEM:",     f"{celkem + dph:,.0f} Kč"],
        ]

        tbl = Table(data_fak, colWidths=[90*mm, 25*mm, 30*mm, 30*mm])
        tbl.setStyle(TableStyle([
            # Záhlaví
            ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#2E86AB")),
            ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN",       (1,0), (-1,-1), "RIGHT"),
            ("FONTSIZE",    (0,0), (-1,-1), 10),
            ("ROWBACKGROUNDS", (0,1), (-1,-4), [colors.white, colors.HexColor("#F2F2F2")]),
            # Shrnutí
            ("FONTNAME",    (2,-3), (-1,-1), "Helvetica-Bold"),
            ("BACKGROUND",  (2,-1), (-1,-1), colors.HexColor("#E8F4FD")),
            ("GRID",        (0,0), (-1,-4), 0.5, colors.grey),
            ("LINEABOVE",   (0,-3), (-1,-3), 1, colors.black),
        ]))
        obsah.append(tbl)
        obsah.append(Spacer(1, 15*mm))

        # Platební podmínky
        obsah.append(Paragraph(
            "Platební podmínky: 14 dní od data vystavení. "
            "Číslo účtu: 1234567890/0100", small))

        doc.build(obsah)

    generuj_fakturu(
        cislo="2024-001",
        odberatel={"nazev": "Firma Novák s.r.o.", "adresa": "Vedlejší 5, Brno", "ico": "98765432"},
        polozky=[
            ("Python kurz – 5 dnů",   1, 15000),
            ("Konzultace (hod)",       3,  2000),
            ("Studijní materiály",     1,  1500),
        ],
        soubor="faktura.pdf",
    )
    print(f"  ✓ faktura.pdf ({Path('faktura.pdf').stat().st_size:,} B)")

    # ── Report se statistikami ────────────────────────────────
    def generuj_report(nazev: str, data: list[dict], soubor: str):
        doc  = SimpleDocTemplate(soubor, pagesize=A4,
                                   leftMargin=20*mm, rightMargin=20*mm,
                                   topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        obsah  = []

        obsah.append(Paragraph(nazev, styles["Title"]))
        obsah.append(Paragraph(
            f"Generováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"]))
        obsah.append(Spacer(1, 8*mm))

        # Tabulka dat
        hlavicky_r = list(data[0].keys())
        tbl_data   = [hlavicky_r] + [[str(r[k]) for k in hlavicky_r] for r in data]
        col_w      = [170*mm // len(hlavicky_r)] * len(hlavicky_r)
        tbl = Table(tbl_data, colWidths=col_w)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#34495E")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("GRID",       (0,0), (-1,-1), 0.3, colors.grey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.white, colors.HexColor("#ECF0F1")]),
        ]))
        obsah.append(tbl)
        doc.build(obsah)

    generuj_report(
        "Zpráva o studentech",
        [{"Jméno": s[1], "Věk": s[2], "Předmět": s[3], "Body": s[4]}
         for s in (studenti if EXCEL_OK else [
             (1, "Míša", 15, "Mat", 87), (2, "Tomáš", 16, "Fyz", 92)])],
        "report.pdf",
    )
    print(f"  ✓ report.pdf ({Path('report.pdf').stat().st_size:,} B)")

    for f in ["faktura.pdf", "report.pdf"]:
        Path(f).unlink(missing_ok=True)


# ══════════════════════════════════════════════════════════════
# ČÁST 3: Čtení existujících PDF (PyMuPDF)
# ══════════════════════════════════════════════════════════════

print("\n=== Čtení PDF (PyMuPDF / pdfplumber) ===\n")
print("""\
# pip install pymupdf
import fitz   # PyMuPDF

doc = fitz.open("dokument.pdf")
for strana in doc:
    text = strana.get_text()
    print(text)
    tabulky = strana.find_tables()   # extrahuj tabulky

# pip install pdfplumber
import pdfplumber
with pdfplumber.open("dokument.pdf") as pdf:
    for strana in pdf.pages:
        text    = strana.extract_text()
        tabulky = strana.extract_tables()
        print(text)
""")

print("""
=== Kdy co použít ===

  openpyxl    → Excel (.xlsx) čtení i zápis, grafy, formátování
  reportlab   → PDF od nuly – faktury, reporty, certifikáty
  PyMuPDF     → čtení a extrakce textu z PDF
  pdfplumber  → extrakce tabulek z PDF
  weasyprint  → HTML → PDF (pip install weasyprint)
  pandas      → data → Excel/CSV (df.to_excel(), df.to_csv())
""")

# TVOJE ÚLOHA:
# 1. Vygeneruj Excel výkaz s podmíněným formátováním (červeně body < 60).
# 2. Vytvoř PDF certifikát pro studenta (jméno, datum, logo).
# 3. Přečti tabulku z PDF faktury a ulož ji do SQLite (lekce 40).
